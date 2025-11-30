"""Django Model Export Functionality

This module provides functionality to export Django model data to Excel or CSV files
through an HTTP endpoint. It supports dynamic model loading, field selection,
filtering, and ordering with GraphQL filter integration.

Features:
- HTTP endpoint for generating downloadable files (JWT protected)
- Support for Excel (.xlsx) and CSV (.csv) formats
- Dynamic model loading by app_name and model_name
- Flexible field selection with nested field access and custom titles
- Advanced filtering using GraphQL filter classes (quick filters, date filters, custom filters)
- Custom ordering support
- Proper error handling and validation

Field Format:
- String format: "field_name" (uses field name as accessor and verbose_name as title)
- Dict format: {"accessor": "field_name", "title": "Custom Title"}

Usage:
    POST /api/export/
    Headers: Authorization: Bearer <jwt_token>
    {
        "app_name": "myapp",
        "model_name": "MyModel",
        "file_extension": "xlsx",
        "filename": "export_data",
        "fields": [
            "title",
            "author.username",
            {"accessor": "slug", "title": "MySlug"}
        ],
        "ordering": ["-created_at"],
        "variables": {
            "status": "active",
            "quick": "search term",
            "published_date_today": true
        }
    }
"""

import csv
import io
import json
import logging
from datetime import date, datetime
from decimal import Decimal
from typing import Any, Dict, List, Optional, Union

from django.apps import apps
from django.core.exceptions import FieldDoesNotExist, ValidationError
from django.db import models
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views import View
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Border, Side

# Import GraphQL filter generator and auth decorators
try:
    from ..generators.filters import AdvancedFilterGenerator
except ImportError:
    AdvancedFilterGenerator = None

try:
    from .auth_decorators import jwt_required
except ImportError:
    jwt_required = None

# Optional Excel support
try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

logger = logging.getLogger(__name__)


class ExportError(Exception):
    """Custom exception for export-related errors."""

    pass


class ModelExporter:
    """
    Handles the export of Django model data to various formats.

    This class provides methods to dynamically load models, apply GraphQL filters,
    extract field data with flexible field format support, and generate export files.

    Features:
    - Dynamic model loading from app and model names
    - GraphQL filter integration for advanced filtering
    - Flexible field format support (string or dict with accessor/title)
    - Nested field access with proper error handling
    - Method calls and property access on model instances
    - Many-to-many field handling
    """

    def __init__(self, app_name: str, model_name: str):
        """
        Initialize the exporter with model information and GraphQL filter generator.

        Args:
            app_name: Name of the Django app containing the model
            model_name: Name of the Django model to export

        Raises:
            ExportError: If the model cannot be found
        """
        self.app_name = app_name
        self.model_name = model_name
        self.model = self._load_model()
        self.logger = logging.getLogger(__name__)

        # Initialize GraphQL filter generator if available
        self.filter_generator = None
        if AdvancedFilterGenerator:
            try:
                self.filter_generator = AdvancedFilterGenerator()
                self.logger.info("GraphQL filter generator initialized successfully")
            except Exception as e:
                self.logger.warning(
                    f"Failed to initialize GraphQL filter generator: {e}"
                )

    def _load_model(self) -> models.Model:
        """
        Load the Django model dynamically.

        Returns:
            The Django model class

        Raises:
            ExportError: If the model cannot be found
        """
        try:
            return apps.get_model(self.app_name, self.model_name)
        except LookupError as e:
            raise ExportError(
                f"Model '{self.model_name}' not found in app '{self.app_name}': {e}"
            )

    def get_queryset(
        self, variables: Optional[Dict[str, Any]] = None, ordering: Optional[str] = None
    ) -> models.QuerySet:
        """
        Get the filtered and ordered queryset using GraphQL filters.

        Args:
            variables: Dictionary of filter kwargs (e.g., {'title__icontains': 'test'})
            ordering: Django ORM ordering expression (e.g., 'created_at' or '-id')

        Returns:
            Filtered and ordered queryset

        Raises:
            ExportError: If filtering or ordering fails
        """
        try:
            queryset = self.model.objects.all()

            # Apply GraphQL filters
            if variables:
                queryset = self.apply_graphql_filters(queryset, variables)

            # Apply ordering
            if ordering:
                queryset = queryset.order_by(ordering)

            return queryset

        except Exception as e:
            raise ExportError(f"Error building queryset: {e}")

    def apply_graphql_filters(
        self, queryset: models.QuerySet, variables: Dict[str, Any]
    ) -> models.QuerySet:
        """
        Apply GraphQL filters to the queryset using the filter generator.

        Args:
            queryset: Django QuerySet to filter
            variables: Filter parameters from the request

        Returns:
            Filtered QuerySet
        """
        if not variables:
            return queryset

        # Try to use GraphQL filter generator first
        if self.filter_generator:
            try:
                # Use apply_complex_filters to handle complex filter structures (AND/OR/NOT)
                # variables likely contains the 'filters' structure from the frontend
                filter_input = variables.get("filters", variables)
                if filter_input is None:
                    return queryset
                return self.filter_generator.apply_complex_filters(
                    queryset, filter_input
                )

            except Exception as e:
                self.logger.warning(
                    f"GraphQL filtering failed, falling back to basic filtering: {e}"
                )

        # Fall back to basic Django filtering
        try:
            # Clean variables to remove None values and empty strings
            clean_variables = {
                key: value
                for key, value in variables.items()
                if value is not None and value != ""
            }
            if clean_variables:
                return queryset.filter(**clean_variables)
            return queryset
        except Exception as e:
            self.logger.error(f"Basic filtering failed: {e}")
            return queryset

    def get_field_value(self, instance: models.Model, accessor: str) -> Any:
        """
        Get field value from model instance using accessor path.

        Supports:
        - Simple fields: 'title'
        - Nested fields: 'author.username'
        - Method calls: 'get_absolute_url'
        - Many-to-many fields: 'tags' (returns comma-separated list)

        Args:
            instance: Model instance
            accessor: Dot-separated path to the field/attribute

        Returns:
            The field value, properly formatted
        """
        try:
            # Split accessor by dots for nested access
            parts = accessor.split(".")
            value = instance

            for part in parts:
                if value is None:
                    return None

                # Handle method calls (if part ends with parentheses)
                if part.endswith("()"):
                    method_name = part[:-2]
                    if hasattr(value, method_name):
                        method = getattr(value, method_name)
                        if callable(method):
                            value = method()
                        else:
                            value = method
                    else:
                        return None
                else:
                    # Regular attribute access
                    if hasattr(value, part):
                        attr = getattr(value, part)

                        # Handle callable attributes (methods)
                        if callable(attr):
                            try:
                                value = attr()
                            except Exception as e:
                                self.logger.debug(f"Method call failed for {part}: {e}")
                                return None
                        else:
                            value = attr
                    else:
                        return None

            # Handle many-to-many relationships
            if hasattr(value, "all"):
                try:
                    items = list(value.all())
                    if items:
                        return ", ".join(str(item) for item in items)
                    return ""
                except Exception:
                    pass

            return self._format_value(value)

        except Exception as e:
            self.logger.warning(
                f"Error accessing field '{accessor}' on {instance}: {e}"
            )
            return None

    def _format_value(self, value: Any) -> Any:
        """
        Format value for export based on its type.

        Args:
            value: The value to format

        Returns:
            Formatted value suitable for export
        """
        if value is None:
            return ""
        elif isinstance(value, bool):
            return "Yes" if value else "No"
        elif isinstance(value, (datetime, date)):
            if isinstance(value, datetime):
                # Convert timezone-aware datetime to local time
                if timezone.is_aware(value):
                    value = timezone.localtime(value)
                return value.strftime("%Y-%m-%d %H:%M:%S")
            else:
                return value.strftime("%Y-%m-%d")
        elif isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, models.Model):
            # For related objects, return string representation
            return str(value)
        elif hasattr(value, "all"):
            # For many-to-many fields, join related objects
            return ", ".join(str(item) for item in value.all())
        else:
            return str(value)

    def parse_field_config(
        self, field_config: Union[str, Dict[str, str]]
    ) -> Dict[str, str]:
        """
        Parse field configuration to extract accessor and title.

        Args:
            field_config: Field configuration (string or dict)

        Returns:
            Dict with 'accessor' and 'title' keys
        """
        if isinstance(field_config, str):
            # String format: use field name as accessor and get verbose name as title
            accessor = field_config
            title = self.get_field_verbose_name(accessor)
            return {"accessor": accessor, "title": title}

        elif isinstance(field_config, dict):
            # Dict format: use provided accessor and title
            accessor = field_config.get("accessor", "")
            title = field_config.get("title", accessor)
            return {"accessor": accessor, "title": title}

        else:
            # Invalid format, use string representation
            accessor = str(field_config)
            title = accessor
            return {"accessor": accessor, "title": title}

    def get_field_verbose_name(self, field_path: str) -> str:
        """
        Get the verbose name for a field path, handling nested fields.

        Args:
            field_path: Field path (e.g., 'title', 'author.username')

        Returns:
            Verbose name of the field
        """
        try:
            parts = field_path.split(".")
            current_model = self.model
            verbose_name = field_path  # Default fallback

            for i, part in enumerate(parts):
                try:
                    field = current_model._meta.get_field(part)

                    if i == len(parts) - 1:  # Last part
                        verbose_name = getattr(field, "verbose_name", part)
                    else:
                        # Navigate to related model
                        if hasattr(field, "related_model"):
                            current_model = field.related_model
                        else:
                            break

                except Exception:
                    # Field not found, use the part name
                    verbose_name = part
                    break

            return str(verbose_name).title()

        except Exception as e:
            self.logger.debug(f"Could not get verbose name for {field_path}: {e}")
            return field_path.replace("_", " ").title()

    def get_field_headers(self, fields: List[Union[str, Dict[str, str]]]) -> List[str]:
        """
        Generate column headers for the export with flexible field format support.

        Args:
            fields: List of field definitions (string or dict format)

        Returns:
            List of column headers
        """
        headers = []

        for field_config in fields:
            parsed_field = self.parse_field_config(field_config)
            headers.append(parsed_field["title"])

        return headers

    def _extract_field_data(self, obj, fields):
        """
        Extract field data from a model instance based on field configurations.

        Args:
            obj: Django model instance
            fields: List of field configurations (string or dict format)

        Returns:
            List of field values for the instance
        """
        row_data = []

        for field_config in fields:
            if isinstance(field_config, str):
                # String format: accessor is the field name
                accessor = field_config
            elif isinstance(field_config, dict):
                # Dict format: get accessor from dict
                accessor = field_config["accessor"]
            else:
                # Invalid format, use empty string
                row_data.append("")
                continue

            try:
                value = self.get_field_value(obj, accessor)
                row_data.append(value)
            except Exception as e:
                # Log error and use empty string as fallback
                logging.getLogger(__name__).warning(
                    f"Error extracting field '{accessor}': {e}"
                )
                row_data.append("")

        return row_data

    def _get_field_headers(self, fields):
        """
        Generate field headers from field configurations.

        Args:
            fields: List of field configurations (string or dict format)

        Returns:
            List of header strings for the export file
        """
        headers = []

        for field_config in fields:
            if isinstance(field_config, str):
                # String format: use verbose_name or field name as title
                accessor = field_config
                title = self._get_verbose_name_for_accessor(accessor)
                headers.append(title)
            elif isinstance(field_config, dict):
                # Dict format: use provided title or fallback to verbose_name
                accessor = field_config["accessor"]
                if "title" in field_config:
                    title = field_config["title"]
                else:
                    title = self._get_verbose_name_for_accessor(accessor)
                headers.append(title)

        return headers

    def _get_verbose_name_for_accessor(self, accessor):
        """
        Get the verbose name for a field accessor, handling nested fields.

        Args:
            accessor: Field accessor string (e.g., 'title', 'author.username')

        Returns:
            String representing the verbose name or field name
        """
        try:
            # Split accessor into parts for nested field access
            parts = accessor.split(".")
            current_model = self.model
            verbose_name = None

            for i, part in enumerate(parts):
                try:
                    field = current_model._meta.get_field(part)

                    if i == len(parts) - 1:
                        # Last part - get verbose name
                        verbose_name = getattr(field, "verbose_name", part)
                        if hasattr(field, "related_model") and field.related_model:
                            # For foreign key fields, might want to include related model info
                            verbose_name = str(verbose_name).title()
                    else:
                        # Intermediate part - move to related model
                        if hasattr(field, "related_model") and field.related_model:
                            current_model = field.related_model
                        else:
                            # Can't traverse further, use remaining parts as fallback
                            remaining_parts = ".".join(parts[i:])
                            verbose_name = remaining_parts.replace("_", " ").title()
                            break

                except FieldDoesNotExist:
                    # Field doesn't exist, might be a method or property
                    # Use the part name as fallback
                    if i == len(parts) - 1:
                        verbose_name = part.replace("_", " ").title()
                    else:
                        # Can't traverse further, use remaining parts as fallback
                        remaining_parts = ".".join(parts[i:])
                        verbose_name = remaining_parts.replace("_", " ").title()
                        break

            return verbose_name or accessor.replace("_", " ").title()

        except Exception:
            # Fallback: use accessor with underscores replaced by spaces
            return accessor.replace("_", " ").title()

    def export_to_csv(
        self,
        fields: List[Union[str, Dict[str, str]]],
        variables: Optional[Dict[str, Any]] = None,
        ordering: Optional[str] = None,
    ) -> str:
        """
        Export model data to CSV format with flexible field format support.

        Args:
            fields: List of field definitions (string or dict format)
            variables: Filter variables
            ordering: Ordering expression

        Returns:
            CSV content as string
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Parse field configurations
        parsed_fields = []
        for field_config in fields:
            parsed_field = self.parse_field_config(field_config)
            parsed_fields.append(parsed_field)

        # Write headers
        headers = [parsed_field["title"] for parsed_field in parsed_fields]
        writer.writerow(headers)

        # Write data rows
        queryset = self.get_queryset(variables, ordering)

        for instance in queryset:
            row = []
            for parsed_field in parsed_fields:
                accessor = parsed_field["accessor"]
                value = self.get_field_value(instance, accessor)
                row.append(value)
            writer.writerow(row)

        return output.getvalue()

    def export_to_excel(
        self,
        fields: List[Union[str, Dict[str, str]]],
        variables: Optional[Dict[str, Any]] = None,
        ordering: Optional[str] = None,
    ) -> bytes:
        """
        Export model data to Excel format with flexible field format support.

        Args:
            fields: List of field definitions (string or dict format)
            variables: Filter variables
            ordering: Ordering expression

        Returns:
            Excel file content as bytes

        Raises:
            ExportError: If openpyxl is not available
        """
        if not EXCEL_AVAILABLE:
            raise ExportError(
                "Excel export requires openpyxl package. Install with: pip install openpyxl"
            )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"{self.model_name} Export"

        # Style definitions
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Parse field configurations
        parsed_fields = []
        for field_config in fields:
            parsed_field = self.parse_field_config(field_config)
            parsed_fields.append(parsed_field)

        # Write headers
        headers = [parsed_field["title"] for parsed_field in parsed_fields]
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin"),
            )

        # Write data rows
        queryset = self.get_queryset(variables, ordering)

        for row_num, instance in enumerate(queryset, 2):
            for col_num, parsed_field in enumerate(parsed_fields, 1):
                accessor = parsed_field["accessor"]
                value = self.get_field_value(instance, accessor)
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                cell.border = Border(
                    left=Side(border_style="thin"),
                    right=Side(border_style="thin"),
                    top=Side(border_style="thin"),
                    bottom=Side(border_style="thin"),
                )
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        workbook.save(output)
        return output.getvalue()


@method_decorator(csrf_exempt, name="dispatch")
class ExportView(View):
    """
    Django view for handling model export requests with JWT authentication.

    Accepts POST requests with JSON payload containing export parameters
    and returns downloadable Excel or CSV files. All requests must include
    a valid JWT token in the Authorization header.

    Authentication:
        Requires JWT token: Authorization: Bearer <token>

    Field Format Examples:
    - String: "title" (uses field name as accessor and verbose_name as title)
    - Dict: {"accessor": "author.username", "title": "Author Name"}

    Filter Examples:
    - Basic: {"status": "active", "is_published": true}
    - Quick search: {"quick": "search term"}
    - Date filters: {"created_date_today": true, "updated_date_this_week": true}
    - Custom filters: {"has_tags": true, "content_length": "medium"}
    """

    @method_decorator(jwt_required if jwt_required else lambda f: f)
    def post(self, request):
        """
        Handle POST request for model export (JWT protected).

        Expected JSON payload:
        {
            "app_name": "blog",
            "model_name": "Post",
            "file_extension": "excel",  // or "csv"
            "filename": "posts_export",  // optional
            "fields": [
                "title",
                "author.username",
                {"accessor": "slug", "title": "MySlug"}
            ],
            "ordering": ["-created_at"],  // optional list
            "variables": {  // optional GraphQL filter parameters
                "status": "active",
                "quick": "search term",
                "published_date_today": true
            }
        }

        Returns:
            HttpResponse with file download or JsonResponse with error
        """
        # Log authenticated user for audit purposes
        if hasattr(request, "user") and request.user.is_authenticated:
            logger.info(
                f"Export request from user: {request.user.username} (ID: {request.user.id})"
            )

        try:
            # Parse JSON payload
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                return JsonResponse({"error": "Invalid JSON payload"}, status=400)

            # Validate required parameters
            required_fields = ["app_name", "model_name", "file_extension", "fields"]
            for field in required_fields:
                if field not in data:
                    return JsonResponse(
                        {"error": f"Missing required field: {field}"}, status=400
                    )

            app_name = data["app_name"]
            model_name = data["model_name"]
            file_extension = data["file_extension"].lower()
            fields = data["fields"]

            # Validate file extension
            if file_extension not in ["xlsx", "csv"]:
                return JsonResponse(
                    {"error": 'file_extension must be "xlsx" or "csv"'}, status=400
                )

            # Validate fields format
            if not isinstance(fields, list) or not fields:
                return JsonResponse(
                    {"error": "fields must be a non-empty list"}, status=400
                )

            # Validate field configurations
            for i, field_config in enumerate(fields):
                if isinstance(field_config, str):
                    continue  # String format is valid
                elif isinstance(field_config, dict):
                    if "accessor" not in field_config:
                        return JsonResponse(
                            {
                                "error": f"Invalid field configuration at index {i}: dict format must contain 'accessor' key"
                            },
                            status=400,
                        )
                else:
                    return JsonResponse(
                        {
                            "error": f"Invalid field configuration at index {i}: field must be string or dict with accessor/title"
                        },
                        status=400,
                    )

            # Optional parameters
            filename = data.get("filename")
            ordering = data.get("ordering", [])
            variables = data.get("variables", {})

            # Generate default filename if not provided
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{model_name}_{timestamp}"

            # Create exporter and generate file
            exporter = ModelExporter(app_name, model_name)

            if file_extension == "xlsx":
                content = exporter.export_to_excel(fields, variables, ordering)
                content_type = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                file_ext = "xlsx"
            else:  # csv
                content = exporter.export_to_csv(fields, variables, ordering)
                content_type = "text/csv"
                file_ext = "csv"

            # Create HTTP response with file download
            response = HttpResponse(content, content_type=content_type)
            response["Content-Disposition"] = (
                f'attachment; filename="{filename}.{file_ext}"'
            )
            response["Content-Length"] = len(content)

            logger.info(
                f"Successfully exported {model_name} data to {file_extension} format"
            )
            return response

        except ExportError as e:
            logger.error(f"Export error: {e}")
            return JsonResponse({"error": str(e)}, status=400)
        except Exception as e:
            logger.error(f"Unexpected error during export: {e}")
            return JsonResponse({"error": "Internal server error"}, status=500)

    @method_decorator(jwt_required if jwt_required else lambda f: f)
    def get(self, request):
        """
        Handle GET request - return API documentation (JWT protected).
        """
        # Log authenticated user for audit purposes
        if hasattr(request, "user") and request.user.is_authenticated:
            logger.info(
                f"Export API documentation request from user: {request.user.username}"
            )

        documentation = {
            "endpoint": "/export",
            "method": "POST",
            "authentication": "JWT token required in Authorization header",
            "description": "Export Django model data to Excel or CSV format with GraphQL filter integration",
            "required_headers": {
                "Authorization": "Bearer <jwt_token>",
                "Content-Type": "application/json",
            },
            "required_parameters": {
                "app_name": "string - Name of the Django app containing the model",
                "model_name": "string - Name of the Django model to export",
                "file_extension": 'string - Either "excel" or "csv"',
                "fields": "array - List of field configurations (string or dict format)",
            },
            "optional_parameters": {
                "filename": "string - Custom filename (default: ModelName_timestamp)",
                "ordering": "array - List of Django ORM ordering expressions",
                "variables": "object - GraphQL filter parameters",
            },
            "field_formats": {
                "string": "Uses field name as accessor and verbose_name as title",
                "dict": "Must contain 'accessor' key, optionally 'title' key",
            },
            "filter_examples": {
                "basic": {"status": "active", "is_published": True},
                "quick_search": {"quick": "search term"},
                "date_filters": {
                    "created_date_today": True,
                    "updated_date_this_week": True,
                },
                "custom_filters": {"has_tags": True, "content_length": "medium"},
            },
            "example_request": {
                "url": "/api/export/",
                "method": "POST",
                "headers": {
                    "Authorization": "Bearer eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9...",
                    "Content-Type": "application/json",
                },
                "payload": {
                    "app_name": "blog",
                    "model_name": "Post",
                    "file_extension": "excel",
                    "filename": "blog_posts_export",
                    "fields": [
                        "title",
                        "author.username",
                        {"accessor": "slug", "title": "MySlug"},
                    ],
                    "ordering": ["-created_at"],
                    "variables": {
                        "status": "active",
                        "quick": "search term",
                        "published_date_today": True,
                    },
                },
            },
            "authentication_errors": {
                "401": "Missing or invalid JWT token",
                "403": "Token valid but insufficient permissions",
            },
        }

        return JsonResponse(documentation, json_dumps_params={"indent": 2})


# URL configuration helper
def get_export_urls():
    """
    Helper function to get URL patterns for the export functionality.

    Usage in urls.py:
        from rail_django_graphql.extensions.exporting import get_export_urls

        urlpatterns = [
            # ... other patterns
        ] + get_export_urls()

    Returns:
        List of URL patterns
    """
    from django.urls import path

    return [
        path("export/", ExportView.as_view(), name="model_export"),
    ]


# Utility functions for programmatic use
def export_model_to_csv(
    app_name: str,
    model_name: str,
    fields: List[Union[str, Dict[str, str]]],
    variables: Optional[Dict[str, Any]] = None,
    ordering: Optional[str] = None,
) -> str:
    """
    Programmatically export model data to CSV format with flexible field format support.

    Args:
        app_name: Name of the Django app
        model_name: Name of the model
        fields: List of field definitions (string or dict format)
        variables: Filter variables
        ordering: Ordering expression

    Returns:
        CSV content as string
    """
    exporter = ModelExporter(app_name, model_name)
    return exporter.export_to_csv(fields, variables, ordering)


def export_model_to_excel(
    app_name: str,
    model_name: str,
    fields: List[Union[str, Dict[str, str]]],
    variables: Optional[Dict[str, Any]] = None,
    ordering: Optional[str] = None,
) -> bytes:
    """
    Programmatically export model data to Excel format with flexible field format support.

    Args:
        app_name: Name of the Django app
        model_name: Name of the model
        fields: List of field definitions (string or dict format)
        variables: Filter variables
        ordering: Ordering expression

    Returns:
        Excel file content as bytes
    """
    exporter = ModelExporter(app_name, model_name)
    return exporter.export_to_excel(fields, variables, ordering)
